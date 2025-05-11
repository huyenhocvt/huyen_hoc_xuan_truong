import os
import io
import json
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
from openpyxl import Workbook, load_workbook

# Load credentials từ biến môi trường
def get_drive_service():
    json_str = os.getenv("GOOGLE_APPLICATION_CREDENTIALS_JSON")
    if not json_str:
        raise ValueError("Thiếu GOOGLE_APPLICATION_CREDENTIALS_JSON")
    creds_info = json.loads(json_str)
    creds = service_account.Credentials.from_service_account_info(creds_info)
    return build("drive", "v3", credentials=creds)

# ID thư mục chứa file xlsx
FOLDER_ID = "1A_VwKn0UqsQIRcbySCzVeSo4guhN0Jdq"

# Liệt kê tất cả file xlsx trong thư mục
def list_xlsx_files():
    service = get_drive_service()
    query = f"'{FOLDER_ID}' in parents and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' and trashed = false"
    results = service.files().list(q=query, fields="files(id, name)").execute()
    return results.get("files", [])

# Ghi thêm dữ liệu vào file theo ID
def write_to_file(file_id, data_rows):
    service = get_drive_service()
    request = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    fh.seek(0)

    wb = load_workbook(fh)
    ws = wb.active
    for row in data_rows:
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    media = MediaIoBaseUpload(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    service.files().update(fileId=file_id, media_body=media).execute()

# Xoá file theo ID
def delete_file(file_id):
    service = get_drive_service()
    service.files().delete(fileId=file_id).execute()

# Tạo file mới trong thư mục
def create_new_file(file_name, header_rows):
    service = get_drive_service()
    wb = Workbook()
    ws = wb.active
    for row in header_rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    metadata = {
        'name': file_name,
        'parents': [FOLDER_ID]
    }
    media = MediaIoBaseUpload(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file = service.files().create(body=metadata, media_body=media, fields='id, name').execute()
    return file
