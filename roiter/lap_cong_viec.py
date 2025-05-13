from flask import Blueprint, render_template, request, session
from datetime import datetime, timedelta
from openpyxl import load_workbook

lap_cong_viec_bp = Blueprint("lap_cong_viec", __name__, url_prefix="/lap-cong-viec")

def get_ten_tu_excel():
    ten_list = []
    try:
        wb = load_workbook("data/danh_sach.xlsx")
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            ten = row[0]  # cột A là 'ten'
            if ten:
                ten_list.append(ten)
    except Exception as e:
        print("Lỗi đọc danh_sach.xlsx:", e)
    return ten_list

@lap_cong_viec_bp.route("/", methods=["GET", "POST"])
def lap_cong_viec():
    if request.method == "GET":
        session["cong_viec_list"] = []

    if request.method == "POST":
        gmt7_now = datetime.utcnow() + timedelta(hours=7)
        ngay_setup = gmt7_now.strftime("%H:%M – %d/%m/%Y")

        try:
            han_hoan_thanh_raw = request.form.get("han_hoan_thanh")
            han_hoan_thanh = datetime.strptime(han_hoan_thanh_raw, "%Y-%m-%dT%H:%M").strftime("%H:%M – %d/%m/%Y")
        except:
            han_hoan_thanh = ""

        row = {
            "ngay_setup": ngay_setup,
            "noi_dung": request.form.get("noi_dung"),
            "han_hoan_thanh": han_hoan_thanh,
            "nguoi_thuc_hien": request.form.get("nguoi_thuc_hien"),
            "loai_viec": request.form.get("loai_viec"),
            "ghi_chu": request.form.get("ghi_chu")
        }

        if "cong_viec_list" not in session:
            session["cong_viec_list"] = []

        session["cong_viec_list"].append(row)
        session.modified = True

    danh_sach_nguoi = get_ten_tu_excel()

    return render_template(
        "lap_cong_viec_day_du.html",
        data=session.get("cong_viec_list", []),
        danh_sach_nguoi=danh_sach_nguoi
    )